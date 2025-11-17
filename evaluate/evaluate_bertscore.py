#!/usr/bin/env python3
"""计算专利问答数据中标准回答与模型回答之间的 BERTScore。"""
from __future__ import annotations

import argparse
from importlib import import_module
from pathlib import Path
from typing import List, Optional, Sequence, Tuple

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path("backend/data/专利问题.xlsx"),
        help="Excel 文件路径 (默认: backend/data/专利问题.xlsx)",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="需要处理的工作表名称，留空则使用首个工作表",
    )
    parser.add_argument(
        "--std-col",
        type=str,
        default="F",
        help="标准回答所在列，默认 F",
    )
    parser.add_argument(
        "--model-col",
        type=str,
        default="K",
        help="模型回答所在列，默认 K",
    )
    parser.add_argument(
        "--output-col",
        type=str,
        default="L",
        help="BERTScore 写入列，默认 L",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="数据起始行 (跳过表头)，默认 2",
    )
    parser.add_argument(
        "--model-type",
        type=str,
        default="bert-base-chinese",
        help="BERTScore 使用的底层模型，默认 bert-base-chinese",
    )
    parser.add_argument(
        "--lang",
        type=str,
        default="zh",
        help="语言代码，默认 zh",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=32,
        help="评分批大小，默认 32",
    )
    parser.add_argument(
        "--device",
        type=str,
        default=None,
        help="运行设备 (如 cuda:0)，留空由 bert-score 自动判断",
    )
    return parser.parse_args()


def normalize_cell(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def collect_pairs(
    worksheet,
    std_col: str,
    model_col: str,
    start_row: int,
) -> Tuple[List[int], List[str], List[str]]:
    row_indices: List[int] = []
    refs: List[str] = []
    hyps: List[str] = []

    for row in range(start_row, worksheet.max_row + 1):
        ref = normalize_cell(worksheet[f"{std_col}{row}"].value)
        hyp = normalize_cell(worksheet[f"{model_col}{row}"].value)
        if ref is None and hyp is None:
            continue
        if ref is None or hyp is None:
            row_indices.append(row)
            refs.append(ref or "")
            hyps.append(hyp or "")
            continue
        row_indices.append(row)
        refs.append(ref)
        hyps.append(hyp)
    return row_indices, refs, hyps


def compute_bertscore(
    refs: Sequence[str],
    hyps: Sequence[str],
    model_type: str,
    lang: str,
    batch_size: int,
    device: Optional[str],
) -> List[float]:
    if not refs:
        return []
    scorer = build_scorer(model_type, lang, batch_size, device)
    _, _, f1 = scorer.score(hyps, refs)
    return [float(score) for score in f1]


def build_scorer(
    model_type: str,
    lang: str,
    batch_size: int,
    device: Optional[str],
):
    try:
        module = import_module("bert_score")
        BERTScorer = getattr(module, "BERTScorer")
    except ImportError as exc:  # pragma: no cover - 运行期依赖提示
        raise SystemExit(
            "未找到 bert-score 库，请先执行 `pip install bert-score` 再重试。"
        ) from exc

    return BERTScorer(
        model_type=model_type,
        lang=lang,
        batch_size=batch_size,
        device=device,
    )


def write_scores(worksheet, rows: Sequence[int], scores: Sequence[float], output_col: str) -> None:
    if not rows:
        return
    header_cell = f"{output_col}1"
    if worksheet[header_cell].value in (None, ""):
        worksheet[header_cell].value = "BERTScore_F1"
    for row, score in zip(rows, scores):
        worksheet[f"{output_col}{row}"].value = round(score, 4)


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook.expanduser().resolve()
    if not workbook_path.exists():
        raise SystemExit(f"未找到 Excel 文件: {workbook_path}")

    workbook = load_workbook(workbook_path)
    worksheet = workbook[args.sheet] if args.sheet else workbook.active

    rows, refs, hyps = collect_pairs(worksheet, args.std_col, args.model_col, args.start_row)
    if not rows:
        raise SystemExit("未在指定列中找到任何文本，终止。")

    scores = compute_bertscore(refs, hyps, args.model_type, args.lang, args.batch_size, args.device)
    write_scores(worksheet, rows, scores, args.output_col)

    workbook.save(workbook_path)
    print(
        f"已为 {len(scores)} 条问答写入 BERTScore (工作表: {worksheet.title}, 文件: {workbook_path})"
    )


if __name__ == "__main__":
    main()
