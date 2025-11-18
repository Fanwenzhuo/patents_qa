import json
import os
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm

from generate import generate_answer
from backend.text_to_sql.text2sql_llm import text2sql


def process_one_row(row_idx, question):
    """在线程池中执行：生成 SQL + 生成答案"""
    sql_query = ""
    answer_content = ""

    # 调用 text2sql
    try:
        sql_query = text2sql(question)
    except Exception as e:
        sql_query = f"错误: {str(e)}"

    # 调用 generate_answer
    try:
        answer_json_str = generate_answer(question)
        answer_data = json.loads(answer_json_str)
        answer_content = answer_data.get("content", "")
    except Exception as e:
        answer_content = f"错误: {str(e)}"

    return {
        "row": row_idx,
        "sql": sql_query,
        "answer": answer_content
    }


def process_excel_concurrent(max_workers=8):
    """并发处理Excel + tqdm 进度条"""

    excel_path = os.path.join("backend", "data", "专利问题.xlsx")
    wb = load_workbook(excel_path)
    ws = wb.active

    total_rows = ws.max_row
    print(f"总共 {total_rows} 行数据需要处理")

    tasks = []
    executor = ThreadPoolExecutor(max_workers=max_workers)

    # 统计可处理行数（非空问题）
    valid_rows = []
    for row_idx in range(2, total_rows + 1):
        question = ws.cell(row=row_idx, column=5).value
        if question and str(question).strip() != "":
            valid_rows.append((row_idx, str(question).strip()))

    total_tasks = len(valid_rows)
    print(f"可处理有效任务数: {total_tasks}")

    # 提交任务
    for row_idx, question in valid_rows:
        future = executor.submit(process_one_row, row_idx, question)
        tasks.append(future)

    # tqdm 主进度条
    with tqdm(total=total_tasks, desc="处理 Excel", ncols=80) as pbar:
        for future in as_completed(tasks):
            res = future.result()
            row_idx = res["row"]

            # 写入 Excel
            ws.cell(row=row_idx, column=10, value=res["sql"])
            ws.cell(row=row_idx, column=11, value=res["answer"])

            wb.save(excel_path)

            pbar.update(1)  # 更新进度条为 +1

    wb.save(excel_path)
    print(f"\n所有数据处理完成，文件已保存到: {excel_path}")


if __name__ == "__main__":
    process_excel_concurrent(max_workers=16)
