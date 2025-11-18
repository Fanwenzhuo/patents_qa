import json
import os
from openpyxl import load_workbook
from generate import generate_answer
from backend.text_to_sql.text2sql_llm import text2sql

def process_excel():
    """处理Excel文件，生成答案并写入表格"""
    # Excel文件路径
    excel_path = os.path.join("backend", "data", "专利问题.xlsx")
    
    # 加载工作簿
    wb = load_workbook(excel_path)
    ws = wb.active  # 获取活动工作表
    
    # 遍历所有行（从第2行开始，假设第1行是表头）
    total_rows = ws.max_row
    print(f"总共 {total_rows} 行数据需要处理")
    
    for row_idx in range(2, total_rows + 1):
        # 读取第E列（索引5）的问题
        question_cell = ws.cell(row=row_idx, column=5)  # E列是第5列
        question = question_cell.value
        
        # 跳过空问题
        if not question or (isinstance(question, str) and question.strip() == ""):
            print(f"第 {row_idx} 行：问题为空，跳过")
            continue
        
        question = str(question).strip()
        print(f"\n处理第 {row_idx} 行问题: {question}")
        
        try:
            # 生成 SQL（写入第J列，索引10）
            sql_query = ""
            try:
                sql_query = text2sql(question)
                print(f"生成的SQL: {sql_query}")
            except Exception as e:
                print(f"生成SQL失败: {str(e)}")
                sql_query = f"错误: {str(e)}"
            
            # 写入第J列
            ws.cell(row=row_idx, column=10, value=sql_query)
            
            # 生成答案（写入第K列，索引11）
            answer_content = ""
            try:
                answer_json_str = generate_answer(question)
                # 解析JSON字符串获取content字段
                answer_data = json.loads(answer_json_str)
                answer_content = answer_data.get("content", "")
                print(f"生成的答案: {answer_content[:100]}...")  # 只打印前100个字符
            except Exception as e:
                print(f"生成答案失败: {str(e)}")
                answer_content = f"错误: {str(e)}"
            
            # 写入第K列
            ws.cell(row=row_idx, column=11, value=answer_content)
            
            # 每处理一行就保存一次，避免数据丢失
            wb.save(excel_path)
            print(f"第 {row_idx} 行处理完成并已保存")
            
        except Exception as e:
            print(f"第 {row_idx} 行处理出错: {str(e)}")
            # 即使出错也尝试保存已处理的数据
            try:
                wb.save(excel_path)
            except:
                pass
    
    # 最终保存
    wb.save(excel_path)
    print(f"\n所有数据处理完成，文件已保存到: {excel_path}")


if __name__ == "__main__":
    process_excel()

